"""
Funciones de procesamiento en background
"""
import os
import json
import datetime
from collections import defaultdict

from app.models import ExtractionRequest, JobStatus
from app.dependencies import (
    jobs, file_handler, termsuite_service, 
    tmx_parser, ollama_translator, excel_exporter, add_ollama_log
)
from app.utils.translation import normalize_translation_options, trim_translation_for_excel


def filter_with_tmx(results: dict, tmx_terms: list) -> dict:
    """Filtrar resultados marcando términos que están en TMX"""
    tmx_set = set(term.lower() for term in tmx_terms)
    
    if "terms" in results:
        for term in results["terms"]:
            term["in_tmx"] = term.get("groupingKey", "").lower() in tmx_set
    
    return results


async def process_extraction(job_id: str, request: ExtractionRequest):
    """Procesar extracción de términos (background task)"""
    try:
        jobs[job_id]["status"] = JobStatus.PROCESSING
        jobs[job_id]["progress"] = 10
        jobs[job_id]["message"] = "Iniciando TermSuite..."
        
        # Ejecutar TermSuite
        corpus_path = file_handler.get_corpus_path(request.corpus_id)
        output_json = file_handler.get_path("outputs", f"{job_id}.json")
        
        jobs[job_id]["progress"] = 30
        jobs[job_id]["message"] = "Extrayendo términos..."
        
        termsuite_service.extract_terms(
            corpus_path=str(corpus_path),
            output_path=str(output_json),
            language=request.language.value,
            min_frequency=request.min_frequency
        )
        
        jobs[job_id]["progress"] = 70
        jobs[job_id]["message"] = "Procesando resultados..."
        
        # Cargar resultados
        with open(output_json, 'r', encoding='utf-8') as f:
            results = json.load(f)
        
        # Filtrar con TMX si se especifica
        if request.use_tmx and request.tmx_id:
            tmx_terms_path = file_handler.get_path("tmx", f"{request.tmx_id}_terms.json")
            with open(tmx_terms_path, 'r', encoding='utf-8') as f:
                tmx_data = json.load(f)
            # Extraer lista de términos (compatible con formato antiguo y nuevo)
            tmx_terms = tmx_data.get('terms', tmx_data) if isinstance(tmx_data, dict) else tmx_data
            results = filter_with_tmx(results, tmx_terms)
        
        jobs[job_id]["progress"] = 90
        jobs[job_id]["message"] = "Generando Excel..."
        
        # Exportar a Excel
        excel_path = file_handler.get_path("outputs", f"{job_id}.xlsx")
        excel_exporter.export(results, str(excel_path))
        
        jobs[job_id]["status"] = JobStatus.COMPLETED
        jobs[job_id]["progress"] = 100
        jobs[job_id]["message"] = "Extracción completada"
        jobs[job_id]["result_file"] = f"{job_id}.xlsx"
        
    except Exception as e:
        jobs[job_id]["status"] = JobStatus.FAILED
        jobs[job_id]["error"] = str(e)
        jobs[job_id]["message"] = f"Error: {str(e)}"


async def process_auto_translations_unified(job_id: str, tmx_id: str, source_lang: str, target_lang: str):
    """
    VERSIÓN UNIFICADA: Procesar traducciones + clasificación en background
    """
    try:
        jobs[job_id]["status"] = JobStatus.PROCESSING
        jobs[job_id]["progress"] = 10
        jobs[job_id]["message"] = "Cargando términos para procesamiento unificado..."
        
        # Cargar términos del TMX
        terms_path = file_handler.get_path("tmx", f"{tmx_id}_terms.json")
        with open(terms_path, 'r', encoding='utf-8') as f:
            terms_data = json.load(f)
        
        terms_list = terms_data.get('terms', [])
        frequencies = terms_data.get('frequencies', {})
        domain_description = terms_data.get('domain_description', '')
        
        jobs[job_id]["progress"] = 20
        jobs[job_id]["message"] = "Obteniendo traducciones TMX..."
        
        # Obtener traducciones TMX existentes
        tmx_dir = file_handler.uploads_dir / 'tmx'
        tmx_file_path = None
        
        if tmx_dir.exists():
            for file in tmx_dir.glob(f"{tmx_id}*"):
                if file.suffix == '.tmx':
                    tmx_file_path = file
                    break
        
        # Procesar traducciones TMX
        tmx_translations = {}
        if tmx_file_path and tmx_file_path.exists():
            try:
                translations = tmx_parser.parse_with_translations(
                    str(tmx_file_path), 
                    source_lang=source_lang,
                    target_lang=target_lang
                )
                
                # Crear índices de traducciones TMX
                trans_dict_exact = defaultdict(set)
                trans_dict_partial = defaultdict(set)
                
                for trans in translations:
                    source = trans.get('source', '').strip()
                    target = trans.get('target', '').strip()
                    if source and target:
                        source_lower = source.lower()
                        trans_dict_exact[source_lower].add(target)
                        for word in source_lower.split():
                            if len(word) > 2:
                                trans_dict_partial[word].add(target)
                
                # Procesar cada término
                for term in terms_list:
                    term_lower = term.lower()
                    
                    # Buscar coincidencia exacta
                    if term_lower in trans_dict_exact:
                        translations_list = list(trans_dict_exact[term_lower])
                        tmx_translations[term] = {
                            'translation': ' | '.join(translations_list),
                            'type': 'Exacto',
                            'variants': len(translations_list)
                        }
                    else:
                        # Buscar coincidencia parcial
                        partial_translations = set()
                        words = term_lower.split()
                        
                        for word in words:
                            if len(word) > 2 and word in trans_dict_partial:
                                partial_translations.update(trans_dict_partial[word])
                        
                        if partial_translations:
                            tmx_translations[term] = {
                                'translation': ' | '.join(list(partial_translations)[:3]),
                                'type': 'Parcial',
                                'variants': len(partial_translations)
                            }
                        else:
                            tmx_translations[term] = {
                                'translation': '',
                                'type': 'No encontrado',
                                'variants': 0
                            }
            
            except Exception as e:
                print(f"Error procesando traducciones TMX: {e}")
        
        jobs[job_id]["progress"] = 40
        jobs[job_id]["message"] = "Preparando términos para procesamiento unificado Ollama..."
        
        # Preparar términos que necesitan procesamiento unificado
        terms_for_unified = []
        for term in terms_list:
            tmx_data = tmx_translations.get(term, {})
            if tmx_data.get('type') in ['Parcial', 'No encontrado']:
                term_data = {
                    'Término': term,
                    'Frecuencia': frequencies.get(term, 1),
                    'Palabras': len(term.split()),
                    'Tipo Match': tmx_data.get('type', 'No encontrado'),
                    'TMX_Context': tmx_data.get('translation', '')
                }
                terms_for_unified.append(term_data)
        
        jobs[job_id]["progress"] = 50
        jobs[job_id]["message"] = f"Procesamiento de {len(terms_for_unified)} términos con Ollama..."
        
        # DECISIÓN: Método unificado vs tradicional
        unified_results = {}
        traditional_translations = {}
        
        if terms_for_unified:
            if domain_description and domain_description.strip():
                # MÉTODO UNIFICADO: Traducción + Clasificación en una sola llamada
                jobs[job_id]["message"] = f"Procesamiento unificado de {len(terms_for_unified)} términos (traducción + clasificación)..."
                add_ollama_log("UNIFIED_BATCH_START", None, "INICIANDO", f"Procesamiento unificado de {len(terms_for_unified)} términos")
                
                unified_results = await ollama_translator.translate_and_classify_unified_batch(
                    terms_for_unified, 
                    source_lang, 
                    target_lang,
                    domain_description,
                    max_concurrent=int(os.getenv('OLLAMA_MAX_CONCURRENT', '10'))
                )
                
                add_ollama_log("UNIFIED_BATCH_COMPLETE", None, "COMPLETADO", f"Procesados {len(unified_results)}/{len(terms_for_unified)} términos")
            else:
                # MÉTODO TRADICIONAL: Solo traducción (sin clasificación de dominio)
                jobs[job_id]["message"] = f"Traduciendo {len(terms_for_unified)} términos con Ollama..."
                add_ollama_log("TRADITIONAL_BATCH_START", None, "INICIANDO", f"Traducción tradicional de {len(terms_for_unified)} términos")
                
                traditional_translations = await ollama_translator.translate_terms_batch(
                    terms_for_unified, 
                    source_lang, 
                    target_lang,
                    max_concurrent=int(os.getenv('OLLAMA_MAX_CONCURRENT', '10'))
                )
                
                add_ollama_log("TRADITIONAL_BATCH_COMPLETE", None, "COMPLETADO", f"Traducidos {len(traditional_translations)}/{len(terms_for_unified)} términos")
        
        jobs[job_id]["progress"] = 80
        jobs[job_id]["message"] = "Generando Excel completo pre-calculado..."
        
        # Crear estructura final para Excel con datos unificados
        excel_data = []
        for idx, term in enumerate(terms_list, 1):
            freq = frequencies.get(term, 1)
            word_count = len(term.split())
            tmx_data = tmx_translations.get(term, {})
            
            item = {
                'Número': idx,
                'Término': term,
                'Frecuencia': freq,
                'Longitud': len(term),
                'Palabras': word_count,
                'Idioma': source_lang,
                'Tipo Match': tmx_data.get('type', 'No encontrado'),
                'Variantes': tmx_data.get('variants', 0)
            }
            
            # Determinar traducción y clasificación final
            if term in unified_results:
                # MÉTODO UNIFICADO: Usar resultado unificado de Ollama
                unified_result = unified_results[term]
                item['Traducción'] = unified_result['translation']
                item['Tipo Match'] = f"{tmx_data.get('type', 'No encontrado')} + Ollama Unificado"
                item['Ollama'] = 'Sí (Unificado)'
                item['Contexto Ollama'] = tmx_data.get('translation', 'Sin contexto TMX')
                
                # Añadir columnas de dominio desde resultado unificado
                item['Relevancia Ámbito'] = unified_result['domain_relevance']
                item['Confianza Ámbito'] = f"{unified_result['confidence']}%"
                item['Razón Ámbito'] = unified_result.get('reason', '')[:200]
            elif term in traditional_translations:
                # MÉTODO TRADICIONAL: Usar solo traducción de Ollama
                traditional_result = traditional_translations[term]
                item['Traducción'] = traditional_result['translation']
                item['Tipo Match'] = f"{tmx_data.get('type', 'No encontrado')} + Ollama"
                item['Ollama'] = 'Sí (Solo traducción)'
                item['Contexto Ollama'] = tmx_data.get('translation', 'Sin contexto TMX')
                
                # Sin clasificación de dominio
                item['Relevancia Ámbito'] = 'No especificado'
                item['Confianza Ámbito'] = 'N/A'
                item['Razón Ámbito'] = 'No se especificó ámbito'
            else:
                # Usar traducción TMX si existe
                item['Traducción'] = tmx_data.get('translation', '')
                item['Ollama'] = 'No necesario' if tmx_data.get('type') == 'Exacto' else 'No disponible'
                item['Contexto Ollama'] = 'No aplicable' if tmx_data.get('type') == 'Exacto' else 'Sin traducción'
                
                # Columnas de dominio para términos no procesados
                if domain_description and domain_description.strip():
                    item['Relevancia Ámbito'] = 'No procesado'
                    item['Confianza Ámbito'] = 'N/A'
                    item['Razón Ámbito'] = 'Término con traducción exacta TMX'
                else:
                    item['Relevancia Ámbito'] = 'No especificado'
                    item['Confianza Ámbito'] = 'N/A'
                    item['Razón Ámbito'] = 'No se especificó ámbito'
            
            excel_data.append(item)
        
        # Normalizar y recortar columna Traducción para Excel
        for item in excel_data:
            if 'Traducción' in item and item.get('Traducción'):
                item['Traducción'] = trim_translation_for_excel(
                    normalize_translation_options(item['Traducción'])
                )
        
        # Guardar datos pre-procesados COMPLETOS
        processed_data_path = file_handler.get_path("tmx", f"{tmx_id}_processed.json")
        with open(processed_data_path, 'w', encoding='utf-8') as f:
            json.dump({
                'tmx_id': tmx_id,
                'source_lang': source_lang,
                'target_lang': target_lang,
                'domain_description': domain_description,
                'total_terms': len(terms_list),
                'unified_processed': len(unified_results),
                'traditional_translated': len(traditional_translations),
                'data': excel_data,
                'processed_at': datetime.datetime.now().isoformat(),
                'processing_type': 'unified' if unified_results else 'traditional'
            }, f, ensure_ascii=False, indent=2)
        
        jobs[job_id]["status"] = JobStatus.COMPLETED
        jobs[job_id]["progress"] = 100
        
        # Mensaje final informativo
        if unified_results:
            jobs[job_id]["message"] = f"Procesamiento unificado completado: {len(unified_results)} términos (traducción + clasificación)"
        elif traditional_translations:
            jobs[job_id]["message"] = f"Traducción completada: {len(traditional_translations)} términos traducidos"
        else:
            jobs[job_id]["message"] = "Procesamiento completado: Solo traducciones TMX utilizadas"
        jobs[job_id]["processed_file"] = f"{tmx_id}_processed.json"
        
    except Exception as e:
        jobs[job_id]["status"] = JobStatus.FAILED
        jobs[job_id]["error"] = str(e)
        jobs[job_id]["message"] = f"Error en procesamiento unificado: {str(e)}"


async def process_tmx_export(export_job_id: str, tmx_id: str, params: dict):
    """Procesar exportación TMX en background con traducción Ollama"""
    try:
        jobs[export_job_id]["status"] = JobStatus.PROCESSING
        jobs[export_job_id]["progress"] = 5
        jobs[export_job_id]["message"] = "Cargando términos..."
        
        # Cargar términos del TMX
        tmx_terms_path = file_handler.get_path("tmx", f"{tmx_id}_terms.json")
        with open(tmx_terms_path, 'r', encoding='utf-8') as f:
            tmx_data = json.load(f)
        
        # Extraer términos (compatible con formato nuevo y antiguo)
        if isinstance(tmx_data, dict):
            terms_list = tmx_data.get('terms', [])
            frequencies = tmx_data.get('frequencies', {})
            language = tmx_data.get('language', 'unknown')
            target_language = tmx_data.get('target_language')
        else:
            terms_list = tmx_data
            frequencies = {}
            language = 'unknown'
            target_language = None
        
        jobs[export_job_id]["progress"] = 15
        jobs[export_job_id]["message"] = "Aplicando filtros..."
        
        # Crear estructura para Excel (misma lógica que endpoint síncrono)
        terms_for_excel = []
        for term in terms_list:
            freq = frequencies.get(term, 1)
            word_count = len(term.split())
            
            # Aplicar filtros
            if params.get('min_frequency') and freq < params['min_frequency']:
                continue
            if params.get('min_words') and word_count < params['min_words']:
                continue
            if params.get('max_words') and word_count > params['max_words']:
                continue
            if params.get('exclude_numbers') and any(char.isdigit() for char in term):
                continue
            if params.get('contains') and params['contains'].lower() not in term.lower():
                continue
            
            terms_for_excel.append({
                'Término': term,
                'Frecuencia': freq,
                'Longitud': len(term),
                'Palabras': word_count,
                'Idioma': language
            })
        
        # Ordenar según parámetros
        sort_by = params.get('sort_by', 'frequency')
        sort_order = params.get('sort_order', 'desc')
        
        if sort_by == "frequency":
            terms_for_excel.sort(key=lambda x: x['Frecuencia'], reverse=(sort_order == "desc"))
        elif sort_by == "alphabetical":
            terms_for_excel.sort(key=lambda x: x['Término'].lower(), reverse=(sort_order == "desc"))
        elif sort_by == "length":
            terms_for_excel.sort(key=lambda x: x['Longitud'], reverse=(sort_order == "desc"))
        elif sort_by == "words":
            terms_for_excel.sort(key=lambda x: x['Palabras'], reverse=(sort_order == "desc"))
        
        # Aplicar top_n después de ordenar
        if params.get('top_n'):
            terms_for_excel = terms_for_excel[:params['top_n']]
        
        # Agregar número después de filtrar y ordenar
        for idx, item in enumerate(terms_for_excel, 1):
            item['Número'] = idx
        
        jobs[export_job_id]["progress"] = 25
        
        # Incluir traducción si se solicita
        if params.get('include_translation') and target_language:
            jobs[export_job_id]["message"] = "Buscando traducciones en TMX..."
            
            # Verificar si ya existe un archivo de traducciones en caché
            translations_cache_path = file_handler.get_path("tmx", f"{tmx_id}_translations.json")
            
            if translations_cache_path.exists():
                # Usar caché
                with open(translations_cache_path, 'r', encoding='utf-8') as f:
                    translations_data = json.load(f)
            else:
                # Buscar el archivo TMX original y procesar traducciones
                tmx_dir = file_handler.uploads_dir / 'tmx'
                tmx_file_path = None
                
                if tmx_dir.exists():
                    for file in tmx_dir.glob(f"{tmx_id}*"):
                        if file.suffix == '.tmx':
                            tmx_file_path = file
                            break
                
                if tmx_file_path and tmx_file_path.exists():
                    try:
                        translations = tmx_parser.parse_with_translations(
                            str(tmx_file_path), 
                            source_lang=language,
                            target_lang=target_language
                        )
                        
                        # Procesar traducciones y crear índices optimizados
                        trans_dict_exact = defaultdict(set)
                        trans_dict_partial = defaultdict(set)
                        
                        for trans in translations:
                            source = trans.get('source', '').strip()
                            target = trans.get('target', '').strip()
                            if source and target:
                                source_lower = source.lower()
                                # Índice exacto
                                trans_dict_exact[source_lower].add(target)
                                # Índice parcial (palabras individuales)
                                for word in source_lower.split():
                                    if len(word) > 2:
                                        trans_dict_partial[word].add(target)
                        
                        # Convertir sets a listas para JSON
                        translations_data = {
                            'exact': {k: list(v) for k, v in trans_dict_exact.items()},
                            'partial': {k: list(v) for k, v in trans_dict_partial.items()}
                        }
                        
                        # Guardar en caché
                        with open(translations_cache_path, 'w', encoding='utf-8') as f:
                            json.dump(translations_data, f, ensure_ascii=False, indent=2)
                    except Exception as e:
                        translations_data = {'exact': {}, 'partial': {}}
                else:
                    translations_data = {'exact': {}, 'partial': {}}
            
            jobs[export_job_id]["progress"] = 40
            
            if translations_data and (translations_data.get('exact') or translations_data.get('partial')):
                trans_dict_exact = translations_data.get('exact', {})
                trans_dict_partial = translations_data.get('partial', {})
                
                # Agregar traducción a cada término
                for item in terms_for_excel:
                    term = item['Término']
                    term_lower = term.lower()
                    
                    # 1. Buscar coincidencia exacta
                    if term_lower in trans_dict_exact:
                        translations_list = trans_dict_exact[term_lower]
                        item['Traducción'] = ' | '.join(translations_list)
                        item['Tipo Match'] = 'Exacto'
                        item['Variantes'] = len(translations_list)
                    else:
                        # 2. Buscar coincidencia parcial
                        partial_translations = set()
                        words = term_lower.split()
                        
                        for word in words:
                            if len(word) > 2 and word in trans_dict_partial:
                                partial_translations.update(trans_dict_partial[word])
                        
                        if partial_translations:
                            item['Traducción'] = ' | '.join(list(partial_translations)[:3])
                            item['Tipo Match'] = 'Parcial'
                            item['Variantes'] = len(partial_translations)
                        else:
                            item['Traducción'] = ''
                            item['Tipo Match'] = 'No encontrado'
                            item['Variantes'] = 0
            
            jobs[export_job_id]["progress"] = 50
            
            # NUEVO: Traducir términos con Match Parcial usando Ollama
            if params.get('use_ollama', True) and ollama_translator.is_available() and target_language:
                jobs[export_job_id]["message"] = "Traduciendo términos con Ollama..."
                
                # Filtrar términos que necesitan traducción con Ollama
                terms_to_translate = [
                    item for item in terms_for_excel 
                    if item.get('Tipo Match') in ['Parcial', 'No encontrado']
                ]
                
                if terms_to_translate:
                    jobs[export_job_id]["message"] = f"Traduciendo {len(terms_to_translate)} términos con Ollama..."
                    add_ollama_log("BATCH_START", None, "INICIANDO", f"Iniciando traducción de {len(terms_to_translate)} términos")
                    
                    # Preparar términos con traducción parcial TMX como contexto
                    terms_with_context = []
                    for item in terms_to_translate:
                        term_data = {
                            'Término': item['Término'],
                            'Frecuencia': item.get('Frecuencia', 1),
                            'Palabras': item.get('Palabras', 1),
                            'Tipo Match': item.get('Tipo Match', 'No encontrado')
                        }
                        
                        # Incluir traducción parcial del TMX como contexto para Ollama
                        tmx_translation = item.get('Traducción', '').strip()
                        if tmx_translation:
                            term_data['TMX_Context'] = tmx_translation
                        
                        terms_with_context.append(term_data)
                    
                    # Traducir en lotes usando el método asíncrono
                    max_concurrent = int(os.getenv('OLLAMA_MAX_CONCURRENT', '10'))
                    ollama_translations = await ollama_translator.translate_terms_batch(
                        terms_with_context, 
                        language, 
                        target_language,
                        max_concurrent=max_concurrent
                    )
                    
                    add_ollama_log("BATCH_COMPLETE", None, "COMPLETADO", f"Traducidos {len(ollama_translations)}/{len(terms_to_translate)} términos")
                    
                    # Actualizar términos con traducciones de Ollama
                    for item in terms_for_excel:
                        if item.get('Tipo Match') in ['Parcial', 'No encontrado']:
                            term = item['Término']
                            if term in ollama_translations:
                                # Obtener resultado completo de Ollama
                                ollama_result = ollama_translations[term]
                                ollama_translation = ollama_result['translation']
                                
                                # REEMPLAZAR completamente la traducción con Ollama (no combinar)
                                item['Traducción'] = ollama_translation
                                
                                # Actualizar tipo de match
                                if item.get('Tipo Match') == 'Parcial':
                                    item['Tipo Match'] = 'Parcial + Ollama'
                                else:
                                    item['Tipo Match'] = 'Ollama'
                                
                                item['Ollama'] = 'Sí'
                                
                                # Contexto Ollama: mostrar la traducción parcial TMX que se usó como contexto
                                tmx_translation = ""
                                for term_ctx in terms_with_context:
                                    if term_ctx['Término'] == term and term_ctx.get('TMX_Context'):
                                        tmx_translation = term_ctx['TMX_Context']
                                        break
                                
                                item['Contexto Ollama'] = tmx_translation if tmx_translation else "Sin contexto TMX"
                            else:
                                item['Ollama'] = 'Error'
                                item['Contexto Ollama'] = 'Error en traducción'
                        else:
                            item['Ollama'] = 'No necesario'
                            item['Contexto Ollama'] = 'No aplicable'
                    
                    jobs[export_job_id]["progress"] = 80
                    jobs[export_job_id]["message"] = f"Traducciones Ollama completadas: {len(ollama_translations)}/{len(terms_to_translate)}"
                else:
                    jobs[export_job_id]["message"] = "No hay términos para traducir con Ollama"
            else:
                if not ollama_translator.is_available():
                    jobs[export_job_id]["message"] = "Ollama no disponible, continuando sin traducciones adicionales..."
                # Agregar columnas Ollama como 'No disponible'
                for item in terms_for_excel:
                    item['Ollama'] = 'No disponible'
                    item['Contexto Ollama'] = 'Servicio no disponible'
        
        # NUEVO: Clasificación de dominio si se especifica
        domain_description = tmx_data.get('domain_description') if isinstance(tmx_data, dict) else None
        if domain_description and domain_description.strip() and ollama_translator.is_available():
            jobs[export_job_id]["progress"] = 82
            jobs[export_job_id]["message"] = "Clasificando términos por relevancia al ámbito..."
            
            # Extraer solo los términos para clasificar
            terms_to_classify = [item['Término'] for item in terms_for_excel]
            
            if terms_to_classify:
                add_ollama_log("DOMAIN_BATCH_START", None, "INICIANDO", f"Iniciando clasificación de {len(terms_to_classify)} términos para ámbito: {domain_description[:50]}...")
                
                # Clasificar términos usando el método asíncrono
                max_concurrent = int(os.getenv('OLLAMA_MAX_CONCURRENT', '10'))
                domain_classifications = await ollama_translator.classify_terms_domain_batch(
                    terms_to_classify, 
                    domain_description, 
                    language,
                    max_concurrent=max_concurrent
                )
                
                add_ollama_log("DOMAIN_BATCH_COMPLETE", None, "COMPLETADO", f"Clasificados {len(domain_classifications)}/{len(terms_to_classify)} términos")
                
                # Actualizar términos con clasificaciones de dominio
                for item in terms_for_excel:
                    term = item['Término']
                    if term in domain_classifications:
                        classification = domain_classifications[term]
                        item['Relevancia Ámbito'] = classification['relevance']
                        item['Confianza Ámbito'] = f"{classification['confidence']}%"
                        item['Razón Ámbito'] = classification.get('reason', '')[:200]
                    else:
                        item['Relevancia Ámbito'] = 'Error'
                        item['Confianza Ámbito'] = '0%'
                        item['Razón Ámbito'] = 'Error en clasificación'
                
                jobs[export_job_id]["message"] = f"Clasificación de ámbito completada: {len(domain_classifications)}/{len(terms_to_classify)}"
            else:
                jobs[export_job_id]["message"] = "No hay términos para clasificar"
        else:
            # Añadir columnas de dominio como 'No disponible' si no se especifica dominio
            for item in terms_for_excel:
                if domain_description and domain_description.strip():
                    item['Relevancia Ámbito'] = 'No disponible'
                    item['Confianza Ámbito'] = 'N/A'
                    item['Razón Ámbito'] = 'Servicio Ollama no disponible'
                else:
                    item['Relevancia Ámbito'] = 'No especificado'
                    item['Confianza Ámbito'] = 'N/A'
                    item['Razón Ámbito'] = 'No se especificó ámbito'
        
        jobs[export_job_id]["progress"] = 85
        jobs[export_job_id]["message"] = "Generando Excel..."
        
        # Normalizar y recortar columna Traducción para Excel
        for item in terms_for_excel:
            if 'Traducción' in item and item.get('Traducción'):
                item['Traducción'] = trim_translation_for_excel(
                    normalize_translation_options(item['Traducción'])
                )
        
        # Generar archivo Excel
        import pandas as pd
        
        # Reordenar columnas para mejor visualización
        preferred_order = ['Número', 'Término', 'Frecuencia', 'Longitud', 'Palabras', 'Idioma', 'Traducción', 'Tipo Match', 'Variantes', 'Ollama', 'Contexto Ollama', 'Relevancia Ámbito', 'Confianza Ámbito', 'Razón Ámbito']
        existing_cols = [col for col in preferred_order if col in (terms_for_excel[0].keys() if terms_for_excel else [])]
        other_cols = [col for col in (terms_for_excel[0].keys() if terms_for_excel else []) if col not in existing_cols]
        
        if terms_for_excel:
            # Reordenar datos
            reordered_data = []
            for item in terms_for_excel:
                reordered_item = {}
                for col in existing_cols + other_cols:
                    reordered_item[col] = item.get(col, '')
                reordered_data.append(reordered_item)
            
            df = pd.DataFrame(reordered_data)
        else:
            df = pd.DataFrame()
        
        # Crear archivo Excel
        excel_filename = f"tmx_{tmx_id}.xlsx"
        excel_path = file_handler.get_path("outputs", excel_filename)
        
        if not df.empty:
            # Usar pandas para escribir Excel más rápido
            with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Términos TMX', index=False)
                
                # Obtener workbook y worksheet para aplicar formato
                workbook = writer.book
                worksheet = writer.sheets['Términos TMX']
                
                # Aplicar formato solo a encabezados
                from openpyxl.styles import Font, PatternFill, Alignment
                
                header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                header_font = Font(color="FFFFFF", bold=True, size=11)
                
                # Formatear solo la primera fila (encabezados)
                for col_idx in range(1, len(df.columns) + 1):
                    cell = worksheet.cell(row=1, column=col_idx)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # Ajustar anchos de columna
                column_widths = {
                    'Número': 10,
                    'Término': 50,
                    'Frecuencia': 12,
                    'Longitud': 12,
                    'Palabras': 12,
                    'Idioma': 12,
                    'Traducción': 80,  # Más ancho para múltiples traducciones
                    'Tipo Match': 20,
                    'Variantes': 12,
                    'Ollama': 15,
                    'Contexto Ollama': 40,
                    'Relevancia Ámbito': 18,
                    'Confianza Ámbito': 15,
                    'Razón Ámbito': 80
                }
                
                for idx, col in enumerate(df.columns, 1):
                    col_letter = chr(64 + idx) if idx <= 26 else chr(64 + idx // 26) + chr(64 + idx % 26)
                    width = column_widths.get(col, 15)
                    worksheet.column_dimensions[col_letter].width = width
                
                # Congelar primera fila
                worksheet.freeze_panes = 'A2'
        else:
            # Crear Excel vacío si no hay datos
            from openpyxl import Workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Términos TMX"
            ws['A1'] = "No se encontraron términos con los filtros aplicados"
            wb.save(excel_path)
        
        jobs[export_job_id]["status"] = JobStatus.COMPLETED
        jobs[export_job_id]["progress"] = 100
        jobs[export_job_id]["message"] = "Exportación completada con traducciones Ollama"
        jobs[export_job_id]["result_file"] = excel_filename
        
    except Exception as e:
        jobs[export_job_id]["status"] = JobStatus.FAILED
        jobs[export_job_id]["error"] = str(e)
        jobs[export_job_id]["message"] = f"Error: {str(e)}"
