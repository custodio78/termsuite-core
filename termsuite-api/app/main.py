"""
Main application file - Refactorizado usando routers modulares
"""
from fastapi import FastAPI, HTTPException, BackgroundTasks
from fastapi.responses import FileResponse, HTMLResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from fastapi.middleware.cors import CORSMiddleware
from typing import Optional
import json
import uuid
import pandas as pd

# Importar routers
from app.routers import web, ollama, tmx, corpus, status
from app.dependencies import (
    jobs, file_handler, ollama_translator
)
from app.models import JobStatus
from app.routers.background_tasks import process_tmx_export
from app.utils.translation import normalize_translation_options, trim_translation_for_excel

app = FastAPI(
    title="LinguaTerms API",
    description="API REST para extracción inteligente de términos técnicos",
    version="1.0.0"
)

# Montar archivos estáticos
app.mount("/static", StaticFiles(directory="app/static"), name="static")

# Templates
templates = Jinja2Templates(directory="app/templates")

# Configurar templates en router web
web.set_templates(templates)

# CORS
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Registrar routers
app.include_router(web.router)
app.include_router(ollama.router)
app.include_router(tmx.router)
app.include_router(corpus.router)
app.include_router(status.router)

# ============================================================================
# ENDPOINTS DE EXPORTACIÓN
# Estos endpoints se mantienen aquí debido a su complejidad y tamaño.
# Se pueden mover a app/routers/export.py en el futuro si es necesario.
# ============================================================================

@app.get("/api/export/tmx-excel-instant/{tmx_id}")
async def export_tmx_instant(tmx_id: str):
    """
    Descarga instantánea de Excel pre-generado con procesamiento unificado
    """
    processed_data_path = file_handler.get_path("tmx", f"{tmx_id}_processed.json")
    
    if processed_data_path.exists():
        with open(processed_data_path, 'r', encoding='utf-8') as f:
            processed_data = json.load(f)
        
        # Verificar que sea procesamiento optimizado (unificado o tradicional mejorado)
        processing_type = processed_data.get('processing_type')
        if processing_type in ['unified', 'traditional']:
            # Generar Excel desde datos pre-procesados
            excel_data = processed_data['data']
            
            # Crear Excel optimizado
            df = pd.DataFrame(excel_data)
            
            excel_filename = f"tmx_unified_{tmx_id}.xlsx"
            excel_path = file_handler.get_path("outputs", excel_filename)
            
            # Generar Excel con formato optimizado
            with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Términos TMX', index=False)
                
                # Aplicar formato
                workbook = writer.book
                worksheet = writer.sheets['Términos TMX']
                
                from openpyxl.styles import Font, PatternFill
                header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                header_font = Font(color="FFFFFF", bold=True, size=11)
                
                for col_idx in range(1, len(df.columns) + 1):
                    cell = worksheet.cell(row=1, column=col_idx)
                    cell.fill = header_fill
                    cell.font = header_font
            
            return FileResponse(
                path=excel_path,
                filename=f"terminos_unificado_{processed_data['source_lang']}_{processed_data['target_lang']}.xlsx",
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
    
    # Fallback al método tradicional si no hay datos unificados
    raise HTTPException(
        status_code=404, 
        detail="Procesamiento unificado no disponible. Use descarga asíncrona."
    )


@app.get("/api/tmx/{tmx_id}/unified-status")
async def check_unified_status(tmx_id: str):
    """
    Verificar estado del procesamiento unificado
    """
    processed_data_path = file_handler.get_path("tmx", f"{tmx_id}_processed.json")
    
    if processed_data_path.exists():
        with open(processed_data_path, 'r', encoding='utf-8') as f:
            processed_data = json.load(f)
        
        processing_type = processed_data.get('processing_type')
        if processing_type in ['unified', 'traditional']:
            return {
                "tmx_id": tmx_id,
                "unified_ready": True,
                "processing_type": processing_type,
                "total_terms": processed_data.get('total_terms', 0),
                "unified_processed": processed_data.get('unified_processed', 0),
                "traditional_translated": processed_data.get('traditional_translated', 0),
                "processed_at": processed_data.get('processed_at'),
                "domain_description": processed_data.get('domain_description'),
                "instant_download_available": True
            }
    
    # Verificar si hay trabajo en progreso
    unified_jobs = [
        job_id for job_id, job in jobs.items() 
        if job.get('type') == 'auto_translation' and job.get('tmx_id') == tmx_id
    ]
    
    if unified_jobs:
        latest_job_id = max(unified_jobs, key=lambda x: jobs.get(x, {}).get('progress', 0))
        latest_job = jobs[latest_job_id]
        return {
            "tmx_id": tmx_id,
            "unified_ready": False,
            "in_progress": True,
            "progress": latest_job.get('progress', 0),
            "message": latest_job.get('message', ''),
            "status": latest_job.get('status')
        }
    
    return {
        "tmx_id": tmx_id,
        "unified_ready": False,
        "in_progress": False,
        "message": "Procesamiento unificado no iniciado"
    }


@app.get("/api/export/excel/{job_id}")
async def export_excel(job_id: str):
    """Exportar resultados a Excel"""
    if job_id not in jobs:
        raise HTTPException(status_code=404, detail="Trabajo no encontrado")
    
    job = jobs[job_id]
    if job["status"] != JobStatus.COMPLETED:
        raise HTTPException(
            status_code=400, 
            detail=f"El trabajo está en estado: {job['status']}"
        )
    
    excel_path = file_handler.get_path("outputs", f"{job_id}.xlsx")
    if not excel_path.exists():
        raise HTTPException(status_code=404, detail="Archivo Excel no encontrado")
    
    return FileResponse(
        path=excel_path,
        filename=f"terms_{job_id}.xlsx",
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


@app.get("/api/export/tmx-excel/{tmx_id}")
async def export_tmx_to_excel(
    tmx_id: str,
    min_frequency: Optional[int] = None,
    top_n: Optional[int] = None,
    min_words: Optional[int] = None,
    max_words: Optional[int] = None,
    sort_by: str = "frequency",
    sort_order: str = "desc",
    format: str = "excel",
    columns: Optional[str] = None,
    exclude_numbers: bool = False,
    contains: Optional[str] = None,
    include_translation: bool = False
):
    """
    Exportar términos de TMX a Excel usando datos pre-procesados (FLUJO MEJORADO)
    
    Args:
        tmx_id: ID del TMX subido previamente
        min_frequency: Frecuencia mínima (ej: 5)
        top_n: Top N términos más frecuentes (ej: 100)
        min_words: Mínimo número de palabras (ej: 2)
        max_words: Máximo número de palabras (ej: 5)
        sort_by: Ordenar por: frequency, alphabetical, length, words
        sort_order: Orden: asc o desc
        format: Formato de salida: excel, csv, json
        columns: Columnas a incluir (separadas por coma)
        exclude_numbers: Excluir términos con números
        contains: Filtrar términos que contengan este texto
        include_translation: Incluir traducción si está disponible
    """
    # Verificar si existen datos pre-procesados (con traducciones)
    processed_data_path = file_handler.get_path("tmx", f"{tmx_id}_processed.json")
    
    if processed_data_path.exists() and include_translation:
        # FLUJO RÁPIDO: Usar datos pre-procesados con traducciones
        with open(processed_data_path, 'r', encoding='utf-8') as f:
            processed_data = json.load(f)
        
        terms_for_excel = processed_data['data'].copy()
        language = processed_data['source_lang']
        
        # Cargar domain_description PRIMERO antes de usarla
        tmx_terms_path = file_handler.get_path("tmx", f"{tmx_id}_terms.json")
        domain_description = None
        if tmx_terms_path.exists():
            with open(tmx_terms_path, 'r', encoding='utf-8') as f:
                tmx_data = json.load(f)
            domain_description = tmx_data.get('domain_description') if isinstance(tmx_data, dict) else None
        
        # NUEVO: Verificar si los datos pre-procesados incluyen columnas de dominio
        # Si no las tienen, añadirlas SOLO si realmente se necesitan
        if (terms_for_excel and 
            'Relevancia Ámbito' not in terms_for_excel[0] and
            domain_description and 
            domain_description.strip() and 
            ollama_translator.is_available()):
            
            # Extraer términos para clasificar
            terms_to_classify = [item['Término'] for item in terms_for_excel]
            
            if terms_to_classify:
                # Clasificar términos
                domain_classifications = await ollama_translator.classify_terms_domain_batch(
                    terms_to_classify, 
                    domain_description, 
                    language,
                    max_concurrent=3
                )
                
                # Añadir columnas de dominio
                for item in terms_for_excel:
                    term = item['Término']
                    if term in domain_classifications:
                        classification = domain_classifications[term]
                        item['Relevancia Ámbito'] = classification['relevance']
                        item['Confianza Ámbito'] = f"{classification['confidence']}%"
                        item['Razón Ámbito'] = classification.get('reason', '')[:200]
                    else:
                        item['Relevancia Ámbito'] = 'Error'
                        item['Confianza Ámbito'] = '0%'
                        item['Razón Ámbito'] = 'Error en clasificación'
            else:
                # Sin términos para clasificar
                for item in terms_for_excel:
                    item['Relevancia Ámbito'] = 'No disponible'
                    item['Confianza Ámbito'] = 'N/A'
                    item['Razón Ámbito'] = 'No hay términos para clasificar'
        elif terms_for_excel and 'Relevancia Ámbito' not in terms_for_excel[0]:
            # Añadir columnas de dominio como 'No especificado' si no se solicitan
            for item in terms_for_excel:
                item['Relevancia Ámbito'] = 'No especificado'
                item['Confianza Ámbito'] = 'N/A'
                item['Razón Ámbito'] = 'No se especificó ámbito'
        
    else:
        # FLUJO TRADICIONAL: Cargar términos básicos sin traducciones
        tmx_terms_path = file_handler.get_path("tmx", f"{tmx_id}_terms.json")
        if not tmx_terms_path.exists():
            raise HTTPException(status_code=404, detail="TMX no encontrado")
        
        with open(tmx_terms_path, 'r', encoding='utf-8') as f:
            tmx_data = json.load(f)
        
        # Extraer términos (compatible con formato nuevo y antiguo)
        if isinstance(tmx_data, dict):
            terms_list = tmx_data.get('terms', [])
            frequencies = tmx_data.get('frequencies', {})
            language = tmx_data.get('language', 'unknown')
        else:
            terms_list = tmx_data
            frequencies = {}
            language = 'unknown'
        
        # Crear estructura básica para Excel
        terms_for_excel = []
        for idx, term in enumerate(terms_list, 1):
            freq = frequencies.get(term, 1)
            word_count = len(term.split())
            
            terms_for_excel.append({
                'Número': idx,
                'Término': term,
                'Frecuencia': freq,
                'Longitud': len(term),
                'Palabras': word_count,
                'Idioma': language,
                'Traducción': '',
                'Tipo Match': 'Sin procesar',
                'Variantes': 0,
                'Ollama': 'No procesado',
                'Contexto Ollama': 'No procesado'
            })
    
    # NUEVO: Añadir clasificación de dominio si hay descripción de ámbito
    # PERO SOLO si no hay muchos términos (para evitar bloquear la descarga)
    # Cargar datos originales para obtener domain_description
    tmx_terms_path = file_handler.get_path("tmx", f"{tmx_id}_terms.json")
    if tmx_terms_path.exists():
        with open(tmx_terms_path, 'r', encoding='utf-8') as f:
            tmx_data = json.load(f)
        domain_description = tmx_data.get('domain_description') if isinstance(tmx_data, dict) else None
    else:
        domain_description = None
    
    if (domain_description and domain_description.strip() and ollama_translator.is_available() 
        and len(terms_for_excel) <= 100):  # LÍMITE: Solo procesar si hay 100 términos o menos
        
        # Extraer solo los términos para clasificar
        terms_to_classify = [item['Término'] for item in terms_for_excel]
        
        if terms_to_classify:
            # Clasificar términos usando el método asíncrono
            domain_classifications = await ollama_translator.classify_terms_domain_batch(
                terms_to_classify, 
                domain_description, 
                language,
                max_concurrent=3  # Limitar concurrencia
            )
            
            # Actualizar términos con clasificaciones de dominio
            for item in terms_for_excel:
                term = item['Término']
                if term in domain_classifications:
                    classification = domain_classifications[term]
                    item['Relevancia Ámbito'] = classification['relevance']
                    item['Confianza Ámbito'] = f"{classification['confidence']}%"
                    item['Razón Ámbito'] = classification.get('reason', '')[:200]
                else:
                    item['Relevancia Ámbito'] = 'Error'
                    item['Confianza Ámbito'] = '0%'
                    item['Razón Ámbito'] = 'Error en clasificación'
        else:
            # Añadir columnas de dominio como 'No disponible' si no hay términos
            for item in terms_for_excel:
                item['Relevancia Ámbito'] = 'No disponible'
                item['Confianza Ámbito'] = 'N/A'
                item['Razón Ámbito'] = 'No hay términos para clasificar'
    else:
        # Añadir columnas de dominio como 'No especificado' o 'Usar descarga asíncrona'
        for item in terms_for_excel:
            if domain_description and domain_description.strip():
                if len(terms_for_excel) > 100:
                    item['Relevancia Ámbito'] = 'Usar descarga asíncrona'
                    item['Confianza Ámbito'] = 'N/A'
                    item['Razón Ámbito'] = 'Demasiados términos para descarga rápida'
                elif not ollama_translator.is_available():
                    item['Relevancia Ámbito'] = 'No disponible'
                    item['Confianza Ámbito'] = 'N/A'
                    item['Razón Ámbito'] = 'Servicio Ollama no disponible'
                else:
                    item['Relevancia Ámbito'] = 'Error'
                    item['Confianza Ámbito'] = 'N/A'
                    item['Razón Ámbito'] = 'Error en configuración'
            else:
                item['Relevancia Ámbito'] = 'No especificado'
                item['Confianza Ámbito'] = 'N/A'
                item['Razón Ámbito'] = 'No se especificó ámbito'
    
    # Aplicar filtros a los datos (pre-procesados o básicos)
    filtered_terms = []
    for item in terms_for_excel:
        term = item['Término']
        freq = item['Frecuencia']
        word_count = item['Palabras']
        
        # Aplicar filtros
        if min_frequency and freq < min_frequency:
            continue
        if min_words and word_count < min_words:
            continue
        if max_words and word_count > max_words:
            continue
        if exclude_numbers and any(char.isdigit() for char in term):
            continue
        if contains and contains.lower() not in term.lower():
            continue
        
        filtered_terms.append(item)
    
    # Ordenar según parámetros
    if sort_by == "frequency":
        filtered_terms.sort(key=lambda x: x['Frecuencia'], reverse=(sort_order == "desc"))
    elif sort_by == "alphabetical":
        filtered_terms.sort(key=lambda x: x['Término'].lower(), reverse=(sort_order == "desc"))
    elif sort_by == "length":
        filtered_terms.sort(key=lambda x: x['Longitud'], reverse=(sort_order == "desc"))
    elif sort_by == "words":
        filtered_terms.sort(key=lambda x: x['Palabras'], reverse=(sort_order == "desc"))
    
    # Aplicar top_n después de ordenar
    if top_n:
        filtered_terms = filtered_terms[:top_n]
    
    # Renumerar después de filtrar y ordenar
    for idx, item in enumerate(filtered_terms, 1):
        item['Número'] = idx
    
    # Si no hay términos después de filtrar
    if not filtered_terms:
        raise HTTPException(
            status_code=404, 
            detail="No se encontraron términos con los filtros aplicados"
        )
    
    terms_for_excel = filtered_terms
    
    # Normalizar y recortar columna Traducción para Excel
    for item in terms_for_excel:
        if 'Traducción' in item and item.get('Traducción'):
            item['Traducción'] = trim_translation_for_excel(
                normalize_translation_options(item['Traducción'])
            )
    
    # Seleccionar columnas si se especifica
    if columns:
        selected_cols = [col.strip().capitalize() for col in columns.split(',')]
        # Mapeo de nombres de columnas
        col_mapping = {
            'Term': 'Término',
            'Frequency': 'Frecuencia',
            'Length': 'Longitud',
            'Words': 'Palabras',
            'Language': 'Idioma',
            'Translation': 'Traducción',
            'Number': 'Número'
        }
        # Convertir nombres en inglés a español
        selected_cols = [col_mapping.get(col, col) for col in selected_cols]
        # Filtrar solo columnas existentes
        available_cols = list(terms_for_excel[0].keys()) if terms_for_excel else []
        selected_cols = [col for col in selected_cols if col in available_cols]
        if selected_cols:
            terms_for_excel = [{k: v for k, v in item.items() if k in selected_cols} 
                              for item in terms_for_excel]
    
    df = pd.DataFrame(terms_for_excel)
    
    # Exportar según formato
    if format == "json":
        output_filename = f"tmx_{tmx_id}.json"
        output_path = file_handler.get_path("outputs", output_filename)
        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(terms_for_excel, f, ensure_ascii=False, indent=2)
        return FileResponse(
            path=output_path,
            filename=f"terminos_tmx_{language}.json",
            media_type="application/json"
        )
    
    elif format == "csv":
        output_filename = f"tmx_{tmx_id}.csv"
        output_path = file_handler.get_path("outputs", output_filename)
        df.to_csv(output_path, index=False, encoding='utf-8-sig')
        return FileResponse(
            path=output_path,
            filename=f"terminos_tmx_{language}.csv",
            media_type="text/csv"
        )
    
    # Formato Excel (por defecto) - OPTIMIZADO
    excel_filename = f"tmx_{tmx_id}.xlsx"
    excel_path = file_handler.get_path("outputs", excel_filename)
    
    # Reordenar columnas para mejor visualización
    preferred_order = ['Número', 'Término', 'Frecuencia', 'Longitud', 'Palabras', 'Idioma', 'Traducción', 'Tipo Match', 'Variantes', 'Ollama', 'Contexto Ollama', 'Relevancia Ámbito', 'Confianza Ámbito', 'Razón Ámbito']
    existing_cols = [col for col in preferred_order if col in df.columns]
    other_cols = [col for col in df.columns if col not in existing_cols]
    df = df[existing_cols + other_cols]
    
    # Usar pandas para escribir Excel más rápido
    with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Términos TMX', index=False)
        
        # Obtener workbook y worksheet para aplicar formato
        workbook = writer.book
        worksheet = writer.sheets['Términos TMX']
        
        # Aplicar formato solo a encabezados (más rápido)
        from openpyxl.styles import Font, PatternFill, Alignment
        
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        
        # Formatear solo la primera fila (encabezados)
        for col_idx in range(1, len(df.columns) + 1):
            cell = worksheet.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Ajustar anchos de columna
        column_widths = {
            'Número': 10,
            'Término': 50,
            'Frecuencia': 12,
            'Longitud': 12,
            'Palabras': 12,
            'Idioma': 12,
            'Traducción': 80,  # Más ancho para traducciones de Ollama
            'Tipo Match': 20,
            'Variantes': 12,
            'Ollama': 15,
            'Contexto Ollama': 60,  # Ancho para contexto TMX
            'Relevancia Ámbito': 18,
            'Confianza Ámbito': 15,
            'Razón Ámbito': 80
        }
        
        for idx, col in enumerate(df.columns, 1):
            col_letter = chr(64 + idx) if idx <= 26 else chr(64 + idx // 26) + chr(64 + idx % 26)
            width = column_widths.get(col, 15)
            worksheet.column_dimensions[col_letter].width = width
        
        # Congelar primera fila
        worksheet.freeze_panes = 'A2'
    
    return FileResponse(
        path=excel_path,
        filename=f"terminos_tmx_{language}.xlsx",
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


@app.post("/api/export/tmx-excel-async/{tmx_id}")
async def export_tmx_to_excel_async(
    tmx_id: str,
    background_tasks: BackgroundTasks,
    min_frequency: Optional[int] = None,
    top_n: Optional[int] = None,
    min_words: Optional[int] = None,
    max_words: Optional[int] = None,
    sort_by: str = "frequency",
    sort_order: str = "desc",
    format: str = "excel",
    columns: Optional[str] = None,
    exclude_numbers: bool = False,
    contains: Optional[str] = None,
    include_translation: bool = False,
    use_ollama: bool = True
):
    """
    Exportar términos de TMX a Excel de forma asíncrona (para archivos grandes)
    """
    export_job_id = str(uuid.uuid4())
    
    # Verificar que existe el TMX
    tmx_terms_path = file_handler.get_path("tmx", f"{tmx_id}_terms.json")
    if not tmx_terms_path.exists():
        raise HTTPException(status_code=404, detail="TMX no encontrado")
    
    # Crear trabajo de exportación
    jobs[export_job_id] = {
        "status": JobStatus.PENDING,
        "progress": 0,
        "message": "Preparando exportación...",
        "type": "export",
        "tmx_id": tmx_id
    }
    
    # Ejecutar en background
    background_tasks.add_task(
        process_tmx_export,
        export_job_id,
        tmx_id,
        {
            "min_frequency": min_frequency,
            "top_n": top_n,
            "min_words": min_words,
            "max_words": max_words,
            "sort_by": sort_by,
            "sort_order": sort_order,
            "format": format,
            "columns": columns,
            "exclude_numbers": exclude_numbers,
            "contains": contains,
            "include_translation": include_translation,
            "use_ollama": use_ollama
        }
    )
    
    return {
        "export_job_id": export_job_id,
        "status": "started",
        "message": "Exportación iniciada en segundo plano"
    }


@app.get("/api/download/export/{export_job_id}")
async def download_export_result(export_job_id: str):
    """Descargar resultado de exportación asíncrona"""
    if export_job_id not in jobs:
        raise HTTPException(status_code=404, detail="Trabajo de exportación no encontrado")
    
    job = jobs[export_job_id]
    if job["status"] != JobStatus.COMPLETED:
        raise HTTPException(
            status_code=400, 
            detail=f"La exportación está en estado: {job['status']}"
        )
    
    result_file = job.get("result_file")
    if not result_file:
        raise HTTPException(status_code=404, detail="Archivo de resultado no encontrado")
    
    file_path = file_handler.get_path("outputs", result_file)
    if not file_path.exists():
        raise HTTPException(status_code=404, detail="Archivo no encontrado en disco")
    
    # Determinar tipo de archivo y nombre
    tmx_id = job.get("tmx_id", "unknown")
    if result_file.endswith('.xlsx'):
        media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        filename = f"terminos_tmx_{tmx_id}.xlsx"
    elif result_file.endswith('.csv'):
        media_type = "text/csv"
        filename = f"terminos_tmx_{tmx_id}.csv"
    else:
        media_type = "application/json"
        filename = f"terminos_tmx_{tmx_id}.json"
    
    return FileResponse(
        path=file_path,
        filename=filename,
        media_type=media_type
    )
