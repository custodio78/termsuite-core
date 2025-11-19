import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from typing import Dict, List
from pathlib import Path


class ExcelExporter:
    """Exportador de términos a Excel con formato"""
    
    def export(self, results: Dict, output_path: str):
        """
        Exportar resultados de TermSuite a Excel
        
        Args:
            results: Diccionario con resultados de TermSuite
            output_path: Ruta del archivo Excel de salida
        """
        # Extraer términos
        terms = results.get('terms', [])
        
        if not terms:
            raise ValueError("No hay términos para exportar")
        
        # Convertir a DataFrame
        df = self._prepare_dataframe(terms)
        
        # Crear Excel con formato
        self._create_formatted_excel(df, output_path)
    
    def _prepare_dataframe(self, terms: List[Dict]) -> pd.DataFrame:
        """Preparar DataFrame desde términos de TermSuite"""
        data = []
        
        for term in terms:
            row = {
                'Término': term.get('groupingKey', ''),
                'Patrón': term.get('pattern', ''),
                'Frecuencia': term.get('frequency', 0),
                'Frec. Documentos': term.get('documentFrequency', 0),
                'Especificidad': round(term.get('specificity', 0), 4),
                'En TMX': 'Sí' if term.get('in_tmx', False) else 'No',
                'Palabras': term.get('words', '')
            }
            data.append(row)
        
        df = pd.DataFrame(data)
        
        # Ordenar por frecuencia descendente
        df = df.sort_values('Frecuencia', ascending=False)
        
        return df
    
    def _create_formatted_excel(self, df: pd.DataFrame, output_path: str):
        """Crear Excel con formato profesional"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Términos Extraídos"
        
        # Estilos
        header_fill = PatternFill(
            start_color="366092",
            end_color="366092",
            fill_type="solid"
        )
        header_font = Font(color="FFFFFF", bold=True, size=11)
        
        # Escribir encabezados
        for col_idx, column in enumerate(df.columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=column)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Escribir datos
        for row_idx, row in enumerate(df.values, 2):
            for col_idx, value in enumerate(row, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = Alignment(vertical='center')
        
        # Ajustar anchos de columna
        column_widths = {
            'A': 40,  # Término
            'B': 20,  # Patrón
            'C': 12,  # Frecuencia
            'D': 18,  # Frec. Documentos
            'E': 15,  # Especificidad
            'F': 10,  # En TMX
            'G': 30   # Palabras
        }
        
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
        
        # Congelar primera fila
        ws.freeze_panes = 'A2'
        
        # Guardar
        wb.save(output_path)
